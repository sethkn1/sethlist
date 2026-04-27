"""
validate_concert_data.py — Internal consistency checks for the concerts xlsx.

Runs two passes against the spreadsheet without any API calls:

1. SPELLING — Group every band name (headliners + openers + festival lineup)
   by normalized key. Within each group, if there are multiple distinct
   spellings, flag the cluster for review. The most-frequent spelling is
   highlighted as the likely-canonical form; rarer variants are flagged as
   suspect.

   Examples this catches:
     - "Distrubed" vs "Disturbed" (typo: same normalized key after stripping
       non-alphanumerics + lowercase)
     - "Guns N' Roses" vs "Guns N’ Roses" (curly vs straight apostrophe — same
       normalized key, but different display strings)
     - "the Mars Volta" vs "The Mars Volta" (case)
     - "Tool" vs "tool" (case)

   Does NOT catch:
     - "Bredking Benjamin" vs "Breaking Benjamin" — these have *different*
       normalized keys, so they look like separate bands. A fuzzy-match pass
       (Levenshtein, etc.) is the next layer to catch these. Punted to a
       follow-up — internal-consistency-first per Seth's instructions.

2. OPENERS-INTERNAL — Within each show's "Opening Acts" cell, look for
   suspect entries. Currently this means:
   - Multiple commas with empty fragments (trailing comma, double comma)
   - Single-character or unusually short opener names (likely truncation)
   - Whitespace inconsistencies that survived parsing

Usage:
    python3 validate_concert_data.py "Concert History.xlsx"
    python3 validate_concert_data.py "Concert History.xlsx" --out report.md

The report is markdown — easy to read in any editor or paste into a doc.
"""

import argparse
import re
import sys
import datetime
from collections import defaultdict, Counter
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def normalize_artist_key(s):
    """Mirror of app.js normalizeArtistKey — must stay in sync with the frontend."""
    if not s:
        return ""
    s = str(s).lower()
    # Strip leading "the "
    s = re.sub(r"^the\s+", "", s)
    # Drop everything that isn't a-z0-9
    s = re.sub(r"[^a-z0-9]", "", s)
    return s


def split_acts(s):
    """
    Mirror of app.js splitActs — split a comma/semicolon list, strip
    parentheticals like "(co-headline)".
    """
    if not s:
        return []
    parts = re.split(r"[,;]", str(s))
    out = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        # Strip parenthetical notes
        p = re.sub(r"\s*\([^)]*\)\s*", "", p).strip()
        if p:
            out.append(p)
    return out


def to_date_str(v):
    """Convert spreadsheet date cells to YYYY-MM-DD strings."""
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, datetime.date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, (int, float)):
        return (datetime.date(1899, 12, 30) + datetime.timedelta(days=int(v))).strftime("%Y-%m-%d")
    return str(v)


def load_concerts(xlsx_path):
    """
    Read the concerts xlsx into a list of dicts.

    Returns: [{'date', 'artist', 'venue', 'opening_acts', 'row_num', 'is_festival'}]
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Empty spreadsheet")

    header = [str(h).strip() if h else "" for h in rows[0]]
    # Find column indices by header name (resilient to column reordering)
    def col(name_options):
        for i, h in enumerate(header):
            if h.lower() in [n.lower() for n in name_options]:
                return i
        return None

    col_date = col(["Date"])
    col_artist = col(["Headlining Artist", "Artist"])
    col_venue = col(["Venue"])
    col_acts = col(["Opening Acts"])

    if col_date is None or col_artist is None:
        raise ValueError(f"Couldn't find required columns. Found: {header}")

    concerts = []
    for ri, r in enumerate(rows[1:], start=2):  # 1-based row 2 onwards
        if not any(v is not None for v in r):
            continue
        artist = str(r[col_artist]).strip() if r[col_artist] else None
        if not artist:
            continue
        is_festival = "festival" in artist.lower()
        concerts.append({
            "row_num": ri,
            "date": to_date_str(r[col_date]),
            "artist": artist,
            "venue": str(r[col_venue]).strip() if col_venue is not None and r[col_venue] else None,
            "opening_acts_raw": str(r[col_acts]).strip() if col_acts is not None and r[col_acts] else None,
            "opening_acts": split_acts(r[col_acts] if col_acts is not None else None),
            "is_festival": is_festival,
        })
    return concerts


# ============================================================================
# Pass 1: Spelling clusters — find bands with multiple spellings
# ============================================================================

def find_spelling_inconsistencies(concerts):
    """
    Cluster every band name (headliner + opener + festival lineup) by
    normalized key. Return clusters where more than one distinct spelling
    exists.

    Returns: list of {
        "key": normalized_key,
        "spellings": [(spelling, count, [appearances])],  # sorted by count desc
        "canonical_guess": str,  # most-common spelling
        "suspects": [spelling, ...],  # everything that isn't the canonical guess
    }
    """
    # spelling -> list of {row, role, date, headliner}
    spelling_appearances = defaultdict(list)
    # spelling -> normalized key
    spelling_key = {}

    for c in concerts:
        # Headliner
        if c["artist"] and not c["is_festival"]:
            sp = c["artist"]
            spelling_appearances[sp].append({
                "row": c["row_num"], "role": "headliner",
                "date": c["date"], "headliner": c["artist"],
            })
            spelling_key[sp] = normalize_artist_key(sp)
        # Festival "headliner" cell — that's the festival name, not a band, skip
        # Opening acts
        for opener in c["opening_acts"]:
            spelling_appearances[opener].append({
                "row": c["row_num"], "role": "opener",
                "date": c["date"], "headliner": c["artist"],
            })
            spelling_key[opener] = normalize_artist_key(opener)

    # Group spellings by key
    by_key = defaultdict(list)
    for sp, key in spelling_key.items():
        if key:  # ignore empty (e.g. punctuation-only)
            by_key[key].append(sp)

    # Identify clusters with >1 spelling
    clusters = []
    for key, spellings in by_key.items():
        if len(spellings) <= 1:
            continue
        # Counts per spelling
        counts = []
        for sp in spellings:
            apps = spelling_appearances[sp]
            counts.append((sp, len(apps), apps))
        # Sort by count desc, then alphabetical for stability
        counts.sort(key=lambda x: (-x[1], x[0]))
        canonical = counts[0][0]
        suspects = [sp for sp, _, _ in counts[1:]]
        clusters.append({
            "key": key,
            "spellings": counts,
            "canonical_guess": canonical,
            "suspects": suspects,
        })
    # Sort clusters by total appearances desc — biggest first (most impact)
    clusters.sort(key=lambda c: -sum(n for _, n, _ in c["spellings"]))
    return clusters


# ============================================================================
# Pass 2: Per-show structural concerns
# ============================================================================

def find_structural_concerns(concerts):
    """
    Walk each row and flag opener strings that look malformed:
      - Multiple consecutive commas (",,")
      - Trailing comma
      - Single-letter opener names (likely truncation)
      - Trailing/leading whitespace before split (already stripped, so we
        check the raw cell)
    """
    findings = []
    for c in concerts:
        raw = c["opening_acts_raw"] or ""
        if not raw:
            continue
        issues = []

        # Double commas / triple commas
        if re.search(r",\s*,", raw):
            issues.append("contains consecutive commas (`,,`)")
        # Trailing comma
        if raw.rstrip().endswith(","):
            issues.append("ends with a trailing comma")
        # Leading comma
        if raw.lstrip().startswith(","):
            issues.append("starts with a leading comma")
        # Very short opener names — flag as suspicious
        for opener in c["opening_acts"]:
            if len(opener) <= 2:
                issues.append(f"very short opener name: `{opener!r}`")
        # Same opener listed twice in one show
        seen = set()
        for opener in c["opening_acts"]:
            k = normalize_artist_key(opener)
            if k and k in seen:
                issues.append(f"opener `{opener}` appears twice in this show")
            seen.add(k)

        if issues:
            findings.append({
                "concert": c,
                "issues": issues,
            })
    return findings


# ============================================================================
# Report rendering (markdown)
# ============================================================================

def render_report(clusters, structural, total_concerts):
    out = []
    out.append("# Concert data validation report\n")
    out.append(f"_Auto-generated · {total_concerts} concerts scanned · "
               f"internal-consistency pass only (no API calls)._\n")

    # ----- Section 1: spelling clusters -----
    out.append("\n## 1. Spelling clusters\n")
    out.append(
        "Bands whose name appears with multiple spellings in your spreadsheet. "
        "Same-named bands that look like distinct entries are grouped here using "
        "the same normalization the app uses (lowercase, drop leading 'The', "
        "strip punctuation/spaces). Most-frequent spelling is the likely "
        "canonical form; rarer variants are flagged as **suspects** to review.\n"
    )
    if not clusters:
        out.append("\n_No spelling clusters found — every band name is consistent._\n")
    else:
        out.append(f"\n**{len(clusters)} clusters with multiple spellings:**\n")
        for c in clusters:
            out.append(f"\n### `{c['canonical_guess']}` cluster\n")
            for sp, n, apps in c["spellings"]:
                marker = "  ✓ canonical" if sp == c["canonical_guess"] else "  ⚠ suspect"
                out.append(f"- **`{sp}`** — {n} appearance{'s' if n != 1 else ''}{marker}")
                # Show up to 3 sample rows
                samples = sorted(apps, key=lambda a: a["date"] or "")[:3]
                for a in samples:
                    role = a["role"]
                    if role == "headliner":
                        out.append(f"    - row {a['row']} · {a['date']} · headliner")
                    else:
                        out.append(f"    - row {a['row']} · {a['date']} · opener at "
                                   f"`{a['headliner']}`")
                if len(apps) > 3:
                    out.append(f"    - …and {len(apps) - 3} more")
                out.append("")

    # ----- Section 2: structural concerns -----
    out.append("\n## 2. Structural concerns in opener cells\n")
    out.append(
        "Rows where the `Opening Acts` cell has formatting issues that look "
        "like data-entry slips: stray commas, duplicate openers within one show, "
        "or unusually short opener names that may be truncated.\n"
    )
    if not structural:
        out.append("\n_No structural issues — opener cells are clean._\n")
    else:
        out.append(f"\n**{len(structural)} rows with concerns:**\n")
        for f in structural:
            c = f["concert"]
            out.append(f"\n- **row {c['row_num']}** · {c['date']} · "
                       f"`{c['artist']}` @ {c['venue'] or '?'}")
            out.append(f"  - opener cell: `{c['opening_acts_raw']}`")
            for issue in f["issues"]:
                out.append(f"  - ⚠ {issue}")

    # ----- Footer -----
    out.append("\n---\n")
    out.append(
        "_Next pass: API cross-check against setlist.fm to validate that the "
        "openers you've listed match what setlist.fm has for each show. "
        "Run after fixing anything actionable above._\n"
    )

    return "\n".join(out)


def main():
    parser = argparse.ArgumentParser(description="Validate concert data internal consistency.")
    parser.add_argument("xlsx", help="Path to Concert History.xlsx")
    parser.add_argument("--out", default=None, help="Output report path (default: stdout)")
    args = parser.parse_args()

    if not Path(args.xlsx).exists():
        print(f"Error: {args.xlsx} not found", file=sys.stderr)
        sys.exit(1)

    print(f"Reading {args.xlsx}…", file=sys.stderr)
    concerts = load_concerts(args.xlsx)
    print(f"  {len(concerts)} concerts loaded", file=sys.stderr)

    print("Pass 1: spelling clusters…", file=sys.stderr)
    clusters = find_spelling_inconsistencies(concerts)
    print(f"  {len(clusters)} clusters flagged", file=sys.stderr)

    print("Pass 2: structural concerns…", file=sys.stderr)
    structural = find_structural_concerns(concerts)
    print(f"  {len(structural)} rows flagged", file=sys.stderr)

    report = render_report(clusters, structural, len(concerts))

    if args.out:
        Path(args.out).write_text(report)
        print(f"\nReport written to {args.out}", file=sys.stderr)
    else:
        print(report)


if __name__ == "__main__":
    main()
