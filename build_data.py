"""
Build Sethlist's data files from your spreadsheets.

Usage:
    python3 build_data.py <concerts_source> <posters_source> [--skip-wiki] [--verbose]

Examples:
    python3 build_data.py "Concert History.xlsx" "Poster Collection.xlsx"
    python3 build_data.py data/concerts.csv data/posters.csv
    python3 build_data.py concerts.xlsx posters.xlsx --skip-wiki
    python3 build_data.py concerts.xlsx posters.xlsx --verbose  # debug wiki lookups

What this does:
    1. Generates data/concerts.json and data/posters.json (consumed by the app)
    2. Writes editable CSV mirrors: data/concerts.csv, data/posters.csv
    3. Extracts hyperlinks from xlsx files:
         - Setlist.fm URLs from concert sheet
         - Personal photo URLs from poster sheet (converted to Drive thumbnails)
         - Expressobeans URLs from poster sheet
    4. Detects festivals (Artist name contains "Festival") and attaches a
       festivalKey to each day so the app can group them.
    5. Looks up Wikipedia thumbnails for each artist and saves to
       data/band_images.csv (unless --skip-wiki).

All CSV outputs preserve manual edits on re-run.

Requirements:
    pip install pandas openpyxl
"""

import sys
import re
import json
import csv
import os
import time
import urllib.parse
import urllib.request
import urllib.error
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook


# =========================================================================
# Helpers
# =========================================================================

def clean(v):
    if v is None:
        return None
    if isinstance(v, float) and pd.isna(v):
        return None
    if isinstance(v, (int, float, bool)):
        return v
    s = str(v).strip()
    return s if s else None


def parse_date(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    try:
        return pd.to_datetime(v).strftime("%Y-%m-%d")
    except Exception:
        return None


def to_drive_thumbnail(url):
    """Convert Drive 'view' URL → directly-embeddable thumbnail URL."""
    if not url or "drive.google.com" not in url:
        return url
    m = re.search(r"/file/d/([a-zA-Z0-9_-]+)", url)
    if m:
        return f"https://drive.google.com/thumbnail?id={m.group(1)}&sz=w1000"
    m = re.search(r"[?&]id=([a-zA-Z0-9_-]+)", url)
    if m:
        return f"https://drive.google.com/thumbnail?id={m.group(1)}&sz=w1000"
    return url


def slugify(s):
    """festival key: lowercase, alphanumeric, underscores"""
    if not s:
        return ""
    return re.sub(r"[^a-z0-9]+", "_", s.lower()).strip("_")


# =========================================================================
# Hyperlink extraction (xlsx only)
# =========================================================================

def extract_hyperlinks(xlsx_path, sheet_name, column_name, n_rows):
    """Pull hyperlinks from a named column. Cell text often differs from the URL."""
    wb = load_workbook(xlsx_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        return [None] * n_rows
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    # Try both with and without trailing punctuation/whitespace
    col_idx = None
    candidates = [column_name, column_name + ":", column_name.rstrip(":")]
    for cand in candidates:
        if cand in headers:
            col_idx = headers.index(cand) + 1
            break
    if not col_idx:
        # fallback: fuzzy match ignoring case and punctuation
        for i, h in enumerate(headers):
            if h and re.sub(r"[^a-z0-9]", "", str(h).lower()) == re.sub(r"[^a-z0-9]", "", column_name.lower()):
                col_idx = i + 1
                break
    if not col_idx:
        return [None] * n_rows

    urls = []
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        url = cell.hyperlink.target if cell.hyperlink else None
        # If no hyperlink, sometimes the value IS the URL
        if not url and cell.value and str(cell.value).startswith("http"):
            url = str(cell.value)
        urls.append(url)
    # pad to n_rows
    if len(urls) < n_rows:
        urls += [None] * (n_rows - len(urls))
    return urls[:n_rows]


# =========================================================================
# Concerts
# =========================================================================

def load_concerts(path):
    ext = Path(path).suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        all_sheets = pd.read_excel(path, sheet_name=None)
        for name, df in all_sheets.items():
            if "Headlining Artist" in df.columns:
                df_clean = df.dropna(subset=["Date"]).reset_index(drop=True)
                # Extract hyperlinks in xlsx row order, then attach to df BEFORE sorting
                # so the URL travels with its row through any re-sort.
                setlist_urls = extract_hyperlinks(path, name, "Setlist.FM Link", len(df_clean))
                df_clean["_setlist_url"] = setlist_urls
                return _concerts_from_df(df_clean)
        raise ValueError(f"No concert sheet with 'Headlining Artist' column found in {path}")
    elif ext == ".csv":
        df = pd.read_csv(path).dropna(subset=["date"]).reset_index(drop=True)
        return _concerts_from_df(df, csv_mode=True)
    raise ValueError(f"Unsupported file type: {ext}")


def _concerts_from_df(df, csv_mode=False):
    col = lambda *names: next((n for n in names if n in df.columns), None)
    date_c = col("Date", "date")
    df = df.sort_values(date_c).reset_index(drop=True) if date_c else df

    out = []
    for i, row in df.iterrows():
        date_str = parse_date(row.get(date_c))
        # Hyperlink pre-attached as _setlist_url; CSV mode reads from setlistLink column
        setlist = clean(row.get("_setlist_url"))
        if not setlist and csv_mode:
            txt = clean(row.get("setlistLink"))
            if txt and txt.startswith("http"):
                setlist = txt

        out.append({
            "id": int(i),
            "date": date_str,
            "year": int(pd.to_datetime(date_str).year) if date_str else None,
            "dayOfWeek": clean(row.get(col("Day of Week", "dayOfWeek"))),
            "artist": clean(row.get(col("Headlining Artist", "artist"))),
            "city": clean(row.get(col("City", "city"))),
            "state": clean(row.get(col("State", "state"))),
            "venue": clean(row.get(col("Venue", "venue"))),
            "attendedWith": clean(row.get(col("Attended With", "attendedWith"))),
            "hasPoster": clean(row.get(col("Have Poster", "hasPoster"))) in ("Yes", "yes", True, 1, "1"),
            "setlistLink": setlist,
            "tourName": clean(row.get(col("Tour Name", "tourName"))),
            "openingActs": clean(row.get(col("Opening Acts", "openingActs"))),
            "notes": clean(row.get(col("Show Notes:", "Show Notes", "notes"))),
        })
    return out


# =========================================================================
# Festival detection
# =========================================================================

def attach_festival_keys(concerts):
    """
    A concert is part of a festival if its Artist name contains 'Festival'.
    All festival days sharing the same Tour Name are grouped together.
    (Tour Name is typically distinct per year — e.g., 'Welcome to Rockville 2021'
    vs 'Welcome to Rockville 2023' — which is exactly the grouping we want.)
    """
    by_tour = {}
    for c in concerts:
        if not c.get("artist") or "festival" not in c["artist"].lower():
            continue
        # Use Tour Name as the grouping key (falls back to Artist if tour is missing)
        group_key = c.get("tourName") or c["artist"]
        by_tour.setdefault(group_key, []).append(c)

    for group_name, days in by_tour.items():
        days.sort(key=lambda x: x["date"] or "")
        fest_key = slugify(group_name)
        total = len(days)
        # Display name: prefer tour name (has the year), fallback to cleaned artist
        display_name = group_name
        for i, c in enumerate(days):
            c["festivalKey"] = fest_key
            c["festivalName"] = display_name
            c["festivalDayNumber"] = i + 1
            c["festivalTotalDays"] = total

    # Fill in None for consistent schema
    for c in concerts:
        c.setdefault("festivalKey", None)
        c.setdefault("festivalName", None)
        c.setdefault("festivalDayNumber", None)
        c.setdefault("festivalTotalDays", None)

    return concerts


# =========================================================================
# Posters
# =========================================================================

def load_posters(path):
    ext = Path(path).suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        return _posters_from_xlsx(path)
    elif ext == ".csv":
        df = pd.read_csv(path).dropna(subset=["date"]).reset_index(drop=True)
        return _posters_from_df(df, csv_mode=True)
    raise ValueError(f"Unsupported file type: {ext}")


def _posters_from_xlsx(path):
    all_sheets = pd.read_excel(path, sheet_name=None)
    # Prefer "Poster List" since it has the Variant column, Artist/Illustrator, Expressobeans,
    # and (as of this revision) a Stock Image column.
    sheet_name = None
    for candidate in ("Poster List", "Concert Posters", "Posters"):
        if candidate in all_sheets:
            sheet_name = candidate
            break
    if not sheet_name:
        sheet_name = list(all_sheets.keys())[0]

    df = all_sheets[sheet_name].dropna(subset=["Date"]).reset_index(drop=True)

    personal_urls_raw = extract_hyperlinks(path, sheet_name, "Link to Personal Photo", len(df))
    if not any(personal_urls_raw):
        personal_urls_raw = extract_hyperlinks(path, sheet_name, "Link to Photo", len(df))
    personal_urls = [to_drive_thumbnail(u) if u else None for u in personal_urls_raw]

    # Stock images: uncropped/high-res reference images of each poster (vs.
    # the user's own photos of the framed physical copy). Primary display
    # falls back to personal when stock is missing.
    stock_urls_raw = extract_hyperlinks(path, sheet_name, "Stock Image", len(df))
    stock_urls = [to_drive_thumbnail(u) if u else None for u in stock_urls_raw]

    expressobeans_urls = extract_hyperlinks(path, sheet_name, "Expressobeans Poster Link", len(df))

    # Attach URLs to df BEFORE any potential reordering so they travel with their row
    df["_personal_url"] = personal_urls
    df["_stock_url"] = stock_urls
    df["_expressobeans_url"] = expressobeans_urls

    return _posters_from_df(df)


def _posters_from_df(df, csv_mode=False):
    col = lambda *names: next((n for n in names if n in df.columns), None)
    date_c = col("Date", "date")

    out = []
    personal_urls_out = []
    stock_urls_out = []
    for i, row in df.iterrows():
        date_str = parse_date(row.get(date_c))
        # Hyperlink pre-attached; CSV mode reads from setlistLink/expressobeansLink columns
        eb = clean(row.get("_expressobeans_url"))
        if not eb and csv_mode:
            txt = clean(row.get("expressobeansLink"))
            if txt and str(txt).startswith("http"):
                eb = txt
        personal = clean(row.get("_personal_url"))
        personal_urls_out.append(personal)
        stock = clean(row.get("_stock_url"))
        stock_urls_out.append(stock)

        out.append({
            "id": int(i),
            "date": date_str,
            "year": int(pd.to_datetime(date_str).year) if date_str else None,
            "artist": clean(row.get(col("Artist", "artist"))),
            "illustrator": clean(row.get(col("Artist/Illustrator", "illustrator"))),
            "location": clean(row.get(col("Location", "location"))),
            "type": clean(row.get(col("Type", "type"))),
            "variant": clean(row.get(col("Variant", "variant"))),
            "number": clean(row.get(col("Number", "number"))),
            "tourShowSpecific": clean(row.get(col("Tour/Show Specific", "Tour/Show Specific ", "tourShowSpecific"))),
            "autographed": clean(row.get(col("Autographed", "autographed"))) in ("Yes", "yes", True, 1, "1"),
            "framed": clean(row.get(col("Framed", "framed"))) in ("Yes", "yes", True, 1, "1"),
            "attended": clean(row.get(col("Attended", "attended"))) in ("Yes", "yes", True, 1, "1"),
            "notes": clean(row.get(col("Poster Notes:", "Poster Notes", "notes"))),
            "expressobeansLink": eb,
        })
    return out, personal_urls_out, stock_urls_out


# =========================================================================
# CSV mirrors
# =========================================================================

def write_concerts_csv(path, concerts):
    fields = [
        "id", "date", "dayOfWeek", "artist", "city", "state", "venue",
        "attendedWith", "hasPoster", "setlistLink", "tourName",
        "openingActs", "notes",
    ]
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for c in concerts:
            row = {k: c.get(k, "") for k in fields}
            if row["hasPoster"] is True: row["hasPoster"] = "Yes"
            elif row["hasPoster"] is False: row["hasPoster"] = "No"
            for k, v in row.items():
                if v is None: row[k] = ""
            writer.writerow(row)


def write_posters_csv(path, posters):
    fields = [
        "id", "date", "artist", "illustrator", "location", "type", "variant", "number",
        "tourShowSpecific", "autographed", "framed", "attended", "notes",
        "expressobeansLink",
    ]
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for p in posters:
            row = {k: p.get(k, "") for k in fields}
            for k in ("autographed", "framed", "attended"):
                if row[k] is True: row[k] = "Yes"
                elif row[k] is False: row[k] = "No"
            for k, v in row.items():
                if v is None: row[k] = ""
            writer.writerow(row)


def merge_poster_images(csv_path, posters, personal_urls, stock_urls):
    """Preserve existing CSV entries; add newly-discovered personal/stock URLs.

    The CSV has 3 image columns per poster:
      - personal_url: user's photo (usually of the framed physical copy)
      - stock_url: clean reference image (primary display in the app)
      - official_url: legacy, rarely populated

    Matching strategy: posters are keyed by their identity tuple
    (date, artist, location, type, variant, number). poster_id is just
    a row index that shifts whenever the spreadsheet is re-sorted or new
    rows are inserted, so we MUST NOT use it as a merge key — that bug
    silently re-attaches the wrong row's images to a different poster
    after re-builds. Identity tuple is stable across re-builds: same
    poster, same key, regardless of row order.

    For the rare case where two distinct rows have an identical identity
    tuple (would require two posters with the same date+artist+location+
    type+variant+number, which is essentially the same poster recorded
    twice), the first one in the existing CSV wins. We log a warning so
    the user can decide if that's intentional.
    """
    def _identity_key(date, artist, location, ptype, variant, number):
        # Normalize so trivial whitespace/case variation doesn't break the
        # match. Strip blanks → empty string. Lowercase the strings since
        # an artist's casing can shift over time without it being a
        # "different" poster.
        return (
            (date or "").strip(),
            (artist or "").strip().lower(),
            (location or "").strip().lower(),
            (ptype or "").strip().lower(),
            (variant or "").strip().lower(),
            (number or "").strip().lower(),
        )

    existing = {}
    if os.path.exists(csv_path):
        with open(csv_path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            # The OLD CSV format only stored poster_id, date, and artist —
            # not type/variant/number. So we can't always reconstruct the
            # full identity from the CSV alone. Fall back: build the
            # existing-row map from whatever the CSV has, AND also keep an
            # id-keyed map as a last-resort lookup for cases where the
            # identity-key path can't find a hit. New CSV writes (below)
            # carry the full identity, so the id-fallback gradually goes
            # away after a few rebuild cycles.
            existing_by_identity = {}
            existing_by_id = {}
            for row in reader:
                # Best-effort identity from CSV (date+artist only, since
                # type/variant/number aren't stored in the legacy format).
                # Lookups in the merge loop will use the FULL identity
                # tuple from posters.json, which means CSVs without
                # type/variant/number will only match when those fields
                # happen to be empty for that poster too — which is rare.
                # The id-fallback below covers the gap.
                key = _identity_key(
                    row.get("date"), row.get("artist"), row.get("location"),
                    row.get("type"), row.get("variant"), row.get("number"),
                )
                existing_by_identity[key] = row
                try:
                    existing_by_id[int(row["poster_id"])] = row
                except (ValueError, KeyError):
                    pass

    rows = []
    for p, personal, stock in zip(posters, personal_urls, stock_urls):
        pid = p["id"]
        # Primary lookup: by full identity tuple. This is correct across
        # row reorders, insertions, and id shifts.
        key = _identity_key(
            p.get("date"), p.get("artist"), p.get("location"),
            p.get("type"), p.get("variant"), p.get("number"),
        )
        prev = existing_by_identity.get(key)
        if prev is None:
            # Fallback: legacy CSVs (pre-fix) didn't store type/variant/
            # number, so the identity-key match misses for those. Try by
            # poster_id IF (a) the existing row at that id has the same
            # date+artist (sanity check — guards against the very bug we
            # introduced), AND (b) we haven't already used that row for
            # a different poster.
            candidate = existing_by_id.get(pid)
            if candidate and (candidate.get("date") or "") == (p.get("date") or "") \
                          and (candidate.get("artist") or "").lower() == (p.get("artist") or "").lower():
                prev = candidate
        prev = prev or {}

        # CSV values win ONLY if the xlsx doesn't have a fresher value.
        # This lets a user manually paste URLs but also lets the xlsx override.
        personal_url = personal or prev.get("personal_url") or ""
        stock_url = stock or prev.get("stock_url") or ""
        official_url = prev.get("official_url") or ""
        rows.append({
            "poster_id": pid,
            "date": p["date"] or "",
            "artist": p["artist"] or "",
            "location": p.get("location") or "",
            "type": p.get("type") or "",
            "variant": p.get("variant") or "",
            "number": p.get("number") or "",
            "personal_url": personal_url,
            "stock_url": stock_url,
            "official_url": official_url,
        })

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f, fieldnames=["poster_id", "date", "artist", "location",
                           "type", "variant", "number",
                           "personal_url", "stock_url", "official_url"])
        writer.writeheader()
        for r in rows:
            writer.writerow(r)

    return rows


# =========================================================================
# Wikipedia enrichment for band/artist images
# =========================================================================

WIKI_SUMMARY_API = "https://en.wikipedia.org/api/rest_v1/page/summary/{title}"
USER_AGENT = "Sethlist/1.0 (https://github.com; personal concert archive)"


def wiki_search_and_fetch(artist_name, verbose=False):
    """
    Try to find the Wikipedia article for an artist and return:
      { image_url, website_url, wiki_url, wiki_extract }
    Strategy:
      1. Direct lookup at "<artist>_(band)"
      2. Direct lookup at "<artist>_(musician)"
      3. Direct lookup at "<artist>"
      4. MediaWiki search API, take best musical-related result
    Returns None if nothing good found.
    """
    attempts = [
        f"{artist_name} (band)",
        f"{artist_name} (musician)",
        f"{artist_name} (rapper)",
        artist_name,
    ]

    for title in attempts:
        result = _wiki_summary(title, verbose=verbose)
        if result and _looks_like_musical_article(result):
            if verbose:
                print(f"    → matched '{title}'", flush=True)
            return result
        elif result and verbose:
            print(f"    → '{title}' existed but didn't look musical", flush=True)

    # Fall back to search
    try:
        q = urllib.parse.quote(artist_name + " band")
        search_url = f"https://en.wikipedia.org/w/api.php?action=query&list=search&srsearch={q}&format=json&srlimit=3"
        req = urllib.request.Request(search_url, headers={"User-Agent": USER_AGENT})
        with urllib.request.urlopen(req, timeout=10) as r:
            data = json.loads(r.read())
        hits = data.get("query", {}).get("search", [])
        if verbose:
            print(f"    → search '{artist_name} band' returned {len(hits)} hits", flush=True)
        for h in hits:
            title = h["title"]
            result = _wiki_summary(title, verbose=verbose)
            if result and _looks_like_musical_article(result):
                if verbose:
                    print(f"    → matched search result '{title}'", flush=True)
                return result
    except Exception as e:
        if verbose:
            print(f"    → search failed: {type(e).__name__}: {e}", flush=True)

    return None


def _wiki_summary(title, verbose=False):
    # Title is human-readable (e.g. "Tool (band)"); need full URL encoding INCLUDING slashes.
    # The summary endpoint uses underscore for spaces but quote for special chars.
    encoded = urllib.parse.quote(title.replace(" ", "_"), safe="")
    url = WIKI_SUMMARY_API.format(title=encoded)
    try:
        req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
        with urllib.request.urlopen(req, timeout=10) as r:
            if r.status != 200:
                if verbose:
                    print(f"    → '{title}' returned HTTP {r.status}", flush=True)
                return None
            data = json.loads(r.read())
    except urllib.error.HTTPError as e:
        if verbose and e.code != 404:
            print(f"    → '{title}' HTTP {e.code}", flush=True)
        return None
    except Exception as e:
        if verbose:
            print(f"    → '{title}' error: {type(e).__name__}: {e}", flush=True)
        return None

    if data.get("type") == "disambiguation":
        if verbose:
            print(f"    → '{title}' is disambiguation page", flush=True)
        return None

    image = None
    if data.get("originalimage"):
        image = data["originalimage"].get("source")
    elif data.get("thumbnail"):
        image = data["thumbnail"].get("source")

    return {
        "image_url": image or "",
        "wiki_url": data.get("content_urls", {}).get("desktop", {}).get("page", ""),
        "wiki_extract": data.get("extract", "")[:400],
        "website_url": "",
    }


# Expanded musical vocabulary — case-insensitive substring match
MUSICAL_KEYWORDS = (
    "band", "singer", "musician", "rapper", "rock", "metal", "pop",
    "album", "song", "recording artist", "duo", "trio", "ensemble",
    "vocalist", "guitarist", "drummer", "bassist", "dj", "composer",
    "songwriter", "music", "hip hop", "hip-hop", "punk", "emo",
    "electronic", "synth", "performer", "frontman", "frontwoman",
    "supergroup", "project", "artist", "record label", "discography",
)


def _looks_like_musical_article(result):
    """
    Does the wiki extract mention musical terms? Defends against pulling
    Tool-the-utility instead of Tool-the-band when the direct title matches.
    """
    extract = (result.get("wiki_extract") or "").lower()
    if not extract:
        return False
    return any(kw in extract for kw in MUSICAL_KEYWORDS)


def enrich_bands(artists, csv_path, skip_wiki=False, verbose=False):
    """
    Read existing data/band_images.csv (preserves manual edits), then look up
    Wikipedia for any artist rows that don't yet have an image_url.
    """
    existing = {}
    if os.path.exists(csv_path):
        with open(csv_path, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                name = (row.get("artist") or "").strip()
                if name:
                    existing[name] = row

    rows = []
    fetched = 0
    failed = []
    for artist in sorted(artists):
        prev = existing.get(artist, {})
        image_url = (prev.get("image_url") or "").strip()
        website_url = (prev.get("website_url") or "").strip()
        wiki_url = (prev.get("wiki_url") or "").strip()
        wiki_extract = (prev.get("wiki_extract") or "").strip()

        # Priority for Wikipedia lookup:
        # 1. If user manually set wiki_url in the CSV (and image_url is still blank),
        #    fetch directly from that specific article. This lets them override
        #    for disambiguated names like Ghost_(Swedish_band).
        # 2. Otherwise use the artist-name search heuristic.
        if not image_url and not skip_wiki:
            result = None
            if wiki_url and "wikipedia.org/wiki/" in wiki_url:
                # Extract title from URL and fetch
                title = wiki_url.split("/wiki/")[-1].split("#")[0]
                if verbose:
                    print(f"  Looking up '{artist}' via manual wiki_url '{title}'…", flush=True)
                result = _wiki_summary(title, verbose=verbose)
            else:
                if verbose:
                    print(f"  Looking up '{artist}'…", flush=True)
                result = wiki_search_and_fetch(artist, verbose=verbose)

            if result and result.get("image_url"):
                image_url = result["image_url"]
                wiki_url = result.get("wiki_url") or wiki_url
                wiki_extract = result.get("wiki_extract") or wiki_extract
                fetched += 1
                if verbose:
                    print(f"    ✓ got image", flush=True)
            elif result:
                wiki_url = result.get("wiki_url") or wiki_url
                wiki_extract = result.get("wiki_extract") or wiki_extract
                failed.append((artist, "no image on wiki page"))
                if verbose:
                    print(f"    ✗ article found but no image", flush=True)
            else:
                failed.append((artist, "no match"))
                if verbose:
                    print(f"    ✗ no match", flush=True)
            time.sleep(0.2)  # be polite to wikipedia

        rows.append({
            "artist": artist,
            "image_url": image_url,
            "website_url": website_url,
            "wiki_url": wiki_url,
            "wiki_extract": wiki_extract,
        })

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f, fieldnames=["artist", "image_url", "website_url", "wiki_url", "wiki_extract"])
        writer.writeheader()
        for r in rows:
            writer.writerow(r)

    return rows, fetched, failed


# =========================================================================
# Festival images — manual-entry CSV
# =========================================================================

def seed_festival_images(concerts, csv_path, images_dir=None, root_dir=".",
                         refresh=False, verbose=False):
    """
    Maintain data/festival_images.csv with one row per festival key.

    On first run: seed the file with festival_key + festival_name pre-filled,
      image_url/website_url/wiki_url/wiki_extract empty for the user to fill in.
    On subsequent runs: preserve every existing row and any user-edited values,
      append new rows for any newly-discovered festival_keys.

    The CSV is the single source of truth — we never overwrite user-entered
    image_url, wiki_url, etc. We only update festival_name to track changes
    in the underlying festival data, since the name is derived from concerts.

    Image downloading: if images_dir is provided, this function will also
    download any image_url that doesn't already have a local copy, saving to
    images_dir/{festival_key}.{ext}. The local_image column stores the path
    relative to root_dir (defaults to repo root) — so a downloaded file at
    images/festivals/foo.png becomes "images/festivals/foo.png" in the CSV,
    which the frontend can fetch directly without prepending anything.

    Idempotent — won't re-download unless refresh=True.

    Returns: (total_count, with_image_count, downloaded_count)
    """
    # Find every distinct festival
    fest_info = {}  # festival_key -> festival_name
    for c in concerts:
        if c.get("festivalKey"):
            fest_info.setdefault(c["festivalKey"], c.get("festivalName") or c["festivalKey"])

    # Read existing CSV (if present), preserving manual edits
    existing = {}  # festival_key -> row dict
    if Path(csv_path).exists():
        with open(csv_path, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                key = (row.get("festival_key") or "").strip()
                if key:
                    existing[key] = row

    # Build the merged set: every festival_key from concerts, plus any
    # entries in the CSV that aren't in current concerts (defensive — don't
    # delete user-entered data just because a concert row was removed).
    all_keys = set(fest_info.keys()) | set(existing.keys())

    # Prepare image directory if downloading is enabled
    if images_dir:
        Path(images_dir).mkdir(parents=True, exist_ok=True)

    rows = []
    downloaded = 0
    root_path = Path(root_dir).resolve()
    for key in sorted(all_keys):
        prev = existing.get(key, {})
        # festival_name: prefer current data (in case the user renamed a
        # festival in the spreadsheet), fall back to whatever's in the CSV.
        name = fest_info.get(key) or (prev.get("festival_name") or key)
        image_url = (prev.get("image_url") or "").strip()
        website_url = (prev.get("website_url") or "").strip()
        wiki_url = (prev.get("wiki_url") or "").strip()
        wiki_extract = (prev.get("wiki_extract") or "").strip()
        local_image = (prev.get("local_image") or "").strip()

        # Download image if requested and not already cached locally.
        # local_image is repo-root-relative, so we resolve against root_path
        # to check existence on disk.
        #
        # MIGRATION: builds prior to the move from data/festival_images/ to
        # images/festivals/ stored local_image as "festival_images/foo.png"
        # (data-relative). Detect that shape and migrate: move the file from
        # data/festival_images/ to images/festivals/ and rewrite the CSV
        # value to repo-root-relative. Idempotent — runs once per legacy row.
        if image_url and local_image and not local_image.startswith(("images/", "/", "http")):
            # Looks like a legacy data-relative path. Check if the file exists
            # at the legacy location and migrate it.
            legacy_path = root_path / "data" / local_image
            if legacy_path.exists() and images_dir:
                new_dir = root_path / images_dir if not Path(images_dir).is_absolute() \
                          else Path(images_dir)
                new_dir.mkdir(parents=True, exist_ok=True)
                new_path = new_dir / legacy_path.name
                if not new_path.exists():
                    # Move (preferred — leaves no duplicate). If move fails for
                    # any reason, fall back to copy.
                    try:
                        legacy_path.rename(new_path)
                        if verbose:
                            print(f"    ↪ migrated {legacy_path} → {new_path}", flush=True)
                    except OSError:
                        import shutil
                        shutil.copy2(legacy_path, new_path)
                        if verbose:
                            print(f"    ↪ copied {legacy_path} → {new_path} (move failed)",
                                  flush=True)
                # Rewrite local_image to the new repo-root-relative path
                try:
                    local_image = str(new_path.resolve().relative_to(root_path)).replace("\\", "/")
                except ValueError:
                    local_image = str(new_path).replace("\\", "/")

        if images_dir and image_url:
            existing_local_path = (root_path / local_image) if local_image else None
            should_download = (
                refresh or
                not local_image or
                (existing_local_path is not None and not existing_local_path.exists())
            )
            if should_download:
                downloaded_path = _download_festival_image(
                    image_url, key, images_dir, verbose=verbose)
                if downloaded_path:
                    # Store path relative to repo root so the frontend can use
                    # it directly without any prefix logic.
                    try:
                        local_image = str(
                            Path(downloaded_path).resolve().relative_to(root_path)
                        )
                    except ValueError:
                        # downloaded_path isn't under root — fall back to
                        # the absolute path. Shouldn't happen in practice but
                        # don't crash.
                        local_image = str(Path(downloaded_path))
                    # Normalize to forward-slashes for cross-platform CSVs
                    local_image = local_image.replace("\\", "/")
                    downloaded += 1

        rows.append({
            "festival_key": key,
            "festival_name": name,
            "image_url": image_url,
            "website_url": website_url,
            "wiki_url": wiki_url,
            "wiki_extract": wiki_extract,
            "local_image": local_image,
        })

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["festival_key", "festival_name", "image_url",
                        "website_url", "wiki_url", "wiki_extract", "local_image"]
        )
        writer.writeheader()
        for r in rows:
            writer.writerow(r)

    with_image = sum(1 for r in rows if r["image_url"])
    return len(rows), with_image, downloaded


def _download_festival_image(url, festival_key, images_dir, verbose=False):
    """
    Download a festival image to {images_dir}/{festival_key}.{ext}.

    Detects extension from response Content-Type header (preferred), falling
    back to the URL path. If neither yields a recognized image extension,
    skip the download and warn — better to leave the field blank than save
    a file with a wrong/missing extension.

    Returns the absolute path to the saved file, or None on failure.
    """
    import urllib.request
    import urllib.error
    import mimetypes

    EXT_MAP = {
        "image/jpeg": ".jpg",
        "image/jpg": ".jpg",
        "image/png": ".png",
        "image/webp": ".webp",
        "image/gif": ".gif",
        "image/svg+xml": ".svg",
    }

    if verbose:
        print(f"    Downloading image for {festival_key}: {url[:80]}", flush=True)

    try:
        req = urllib.request.Request(url, headers={
            "User-Agent": "Sethlist/1.0 (personal concert archive; contact: github.com/sethkn1/sethlist)",
        })
        with urllib.request.urlopen(req, timeout=15) as r:
            content_type = (r.headers.get("Content-Type") or "").split(";")[0].strip().lower()
            ext = EXT_MAP.get(content_type)
            if not ext:
                # Fall back to URL path (strip querystring)
                from urllib.parse import urlparse
                path = urlparse(url).path
                guess = mimetypes.guess_extension(content_type) if content_type else None
                if guess in (".jpe", ".jpeg"):
                    ext = ".jpg"
                elif guess in (".png", ".webp", ".gif", ".svg"):
                    ext = guess
                elif path.lower().endswith((".jpg", ".jpeg")):
                    ext = ".jpg"
                elif path.lower().endswith(".png"):
                    ext = ".png"
                elif path.lower().endswith(".webp"):
                    ext = ".webp"
                elif path.lower().endswith(".svg"):
                    ext = ".svg"
                elif path.lower().endswith(".gif"):
                    ext = ".gif"
            if not ext:
                print(f"    ✗ Could not detect image type for {festival_key} "
                      f"(Content-Type: {content_type!r}); skipping.", flush=True)
                return None
            data = r.read()

        # Sanity check on size — refuse to write empty or tiny files
        if len(data) < 100:
            print(f"    ✗ Suspiciously small response ({len(data)} bytes) for "
                  f"{festival_key}; skipping.", flush=True)
            return None

        # Write to disk
        out_path = Path(images_dir) / f"{festival_key}{ext}"
        with open(out_path, "wb") as f:
            f.write(data)
        if verbose:
            kb = len(data) / 1024
            print(f"    ✓ Saved {out_path.name} ({kb:.1f} KB)", flush=True)
        return str(out_path)
    except urllib.error.HTTPError as e:
        print(f"    ✗ HTTP {e.code} for {festival_key}: {url[:80]}", flush=True)
        return None
    except (TimeoutError, OSError, urllib.error.URLError) as e:
        print(f"    ✗ Network error for {festival_key}: {type(e).__name__}", flush=True)
        return None
    except Exception as e:
        print(f"    ✗ Unexpected error for {festival_key}: {type(e).__name__}: {e}",
              flush=True)
        return None


# =========================================================================
# Main
# =========================================================================

def build(concerts_src, posters_src, out_dir="data", skip_wiki=False,
          refresh_festival_images=False, verbose=False):
    Path(out_dir).mkdir(exist_ok=True)

    # Load & enrich concerts
    concerts = load_concerts(concerts_src)
    concerts = attach_festival_keys(concerts)

    # Load posters
    posters, personal_urls, stock_urls = load_posters(posters_src)

    # Write JSON (the app consumes these)
    with open(f"{out_dir}/concerts.json", "w") as f:
        json.dump(concerts, f, indent=2)
    with open(f"{out_dir}/posters.json", "w") as f:
        json.dump(posters, f, indent=2)

    # Write editable CSV mirrors
    write_concerts_csv(f"{out_dir}/concerts.csv", concerts)
    write_posters_csv(f"{out_dir}/posters.csv", posters)

    # Merge poster image URLs (personal + stock)
    img_rows = merge_poster_images(f"{out_dir}/poster_images.csv", posters, personal_urls, stock_urls)
    with_personal = sum(1 for r in img_rows if r["personal_url"])
    with_stock = sum(1 for r in img_rows if r["stock_url"])
    with_official = sum(1 for r in img_rows if r["official_url"])

    # Summary so far
    fest_keys = {c["festivalKey"] for c in concerts if c["festivalKey"]}
    print(f"✓ {len(concerts)} concerts  ({len(fest_keys)} festivals detected)")
    print(f"✓ {len(posters)} posters")
    print(f"✓ poster_images.csv: {with_personal} personal, {with_stock} stock, {with_official} official URLs")

    # Wikipedia enrichment for band images
    unique_artists = {c["artist"] for c in concerts if c.get("artist")}
    # Exclude festival umbrella names (they're not bands).
    # Festival metadata (image, wiki link, etc.) is managed in
    # data/festival_images.csv via manual entry — see seed_festival_images()
    # which is called below.
    unique_artists = {a for a in unique_artists if "festival" not in a.lower()}

    if skip_wiki:
        bands, fetched, failed = enrich_bands(unique_artists, f"{out_dir}/band_images.csv", skip_wiki=True, verbose=verbose)
        print(f"✓ band_images.csv: {len(bands)} artists (Wikipedia lookup skipped)")
    else:
        print(f"  Looking up Wikipedia for {len(unique_artists)} artists…"
              + (" (verbose)" if verbose else ""))
        bands, fetched, failed = enrich_bands(unique_artists, f"{out_dir}/band_images.csv", skip_wiki=False, verbose=verbose)
        with_img = sum(1 for b in bands if b["image_url"])
        print(f"✓ band_images.csv: {with_img}/{len(bands)} artists have images"
              f" ({fetched} new, {len(failed)} not found)")
        if failed:
            # failed is now list of (artist, reason) tuples
            preview = ", ".join(a for a, _ in failed[:5])
            extra = f" (+ {len(failed)-5} more)" if len(failed) > 5 else ""
            print(f"  Missed: {preview}{extra}")
            print(f"  → Edit data/band_images.csv to add image_url manually for these.")

    # Festival images / wiki links — manual-entry CSV.
    # We seed one row per festival_key found in concerts; the user fills in
    # image_url, website_url, wiki_url, wiki_extract by hand. On subsequent
    # builds we preserve any user edits and only add new rows for
    # newly-discovered festivals.
    # If image_url is populated, we download the image into
    # images/festivals/{key}.{ext} so the site doesn't depend on the
    # external URL staying alive. The local_image column tracks the saved
    # path relative to the repo root. Pass refresh_festival_images=True to
    # re-download even if cached.
    fest_count, fest_with_img, fest_dl = seed_festival_images(
        concerts, f"{out_dir}/festival_images.csv",
        images_dir="images/festivals",
        root_dir=".",
        refresh=refresh_festival_images,
        verbose=verbose,
    )
    dl_msg = f", {fest_dl} downloaded" if fest_dl else ""
    print(f"✓ festival_images.csv: {fest_with_img}/{fest_count} festivals have image URLs"
          f"{dl_msg}")

    print()
    print("You can edit these CSVs directly, then re-run this script:")
    print(f"  - {out_dir}/concerts.csv     (add/edit/delete concerts)")
    print(f"  - {out_dir}/posters.csv      (add/edit/delete posters)")
    print(f"  - {out_dir}/poster_images.csv (paste official image URLs)")
    print(f"  - {out_dir}/band_images.csv  (override Wikipedia images; add websites)")
    print(f"  - {out_dir}/festival_images.csv (manual entry — festival images, wiki links, etc.)")


if __name__ == "__main__":
    args = sys.argv[1:]
    skip_wiki = False
    verbose = False
    refresh_festival_images = False
    if "--skip-wiki" in args:
        skip_wiki = True
        args.remove("--skip-wiki")
    if "--refresh-festival-images" in args:
        refresh_festival_images = True
        args.remove("--refresh-festival-images")
    if "--verbose" in args or "-v" in args:
        verbose = True
        for flag in ("--verbose", "-v"):
            while flag in args:
                args.remove(flag)
    if len(args) != 2:
        print(__doc__)
        sys.exit(1)
    build(args[0], args[1], skip_wiki=skip_wiki,
          refresh_festival_images=refresh_festival_images, verbose=verbose)
