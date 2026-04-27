"""
validate_openers_api.py — Cross-check opener bands against setlist.fm.

For each non-festival headliner show in your spreadsheet, query setlist.fm for
all bands that played the same city on the same date, then compare that to the
"Opening Acts" cell in your data. Generate a markdown report flagging:

  - Openers in your sheet that setlist.fm doesn't show (possible misspelling
    or the band actually didn't play that night)
  - Bands setlist.fm shows that aren't in your sheet (you may have missed them
    or skipped them — informational, not always an error)
  - Shows where the API confirms your list matches exactly (clean ✓)
  - Shows where the API has no data at all (informational, not actionable)

Why we don't validate festivals:
  - Festival "openers" in your spreadsheet are the bands YOU saw, not the
    full lineup. Setlist.fm has every band that played across all stages,
    which can be 40+ acts. Reporting "your list is a subset" isn't useful —
    you'd be flagged for everything.

Why we use date+city, not date+venue:
  - The cached setlists in data/setlists.json don't carry venueId (it's
    stripped by prefetch_setlists.py's simplify()). Searching by city is
    almost as precise — most cities have at most one show on a given date —
    and it works for festival venues where the venue name might differ
    between sources.

Usage:
    python3 validate_openers_api.py "Concert History.xlsx"
    python3 validate_openers_api.py "Concert History.xlsx" --out report.md
    python3 validate_openers_api.py "Concert History.xlsx" --limit 5  # smoke test
    python3 validate_openers_api.py "Concert History.xlsx" --refresh  # re-query

Reads:
  - The xlsx for ground truth on what YOU said the openers were
  - data/setlists.json for the headliner artist's canonical name
  - .secrets / SETLIST_FM_KEY for the API key

Caches API responses to data/opener_validation_cache.json so re-runs are fast
(only re-queries if you pass --refresh).
"""

import argparse
import datetime
import json
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


API_BASE = "https://api.setlist.fm/rest/1.0"


# ============================================================================
# Shared utilities (same logic as validate_concert_data.py + prefetch_setlists.py)
# ============================================================================

def normalize_artist_key(s):
    """Mirror of app.js normalizeArtistKey."""
    if not s:
        return ""
    s = str(s).lower()
    s = re.sub(r"^the\s+", "", s)
    s = re.sub(r"[^a-z0-9]", "", s)
    return s


def split_acts(s):
    """Mirror of app.js splitActs."""
    if not s:
        return []
    parts = re.split(r"[,;]", str(s))
    out = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        p = re.sub(r"\s*\([^)]*\)\s*", "", p).strip()
        if p:
            out.append(p)
    return out


def to_date_str(v):
    """Spreadsheet cell to YYYY-MM-DD."""
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, datetime.date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, (int, float)):
        return (datetime.date(1899, 12, 30) + datetime.timedelta(days=int(v))).strftime("%Y-%m-%d")
    return str(v)


def iso_to_setlistfm_date(iso):
    """YYYY-MM-DD → DD-MM-YYYY (setlist.fm's expected format)."""
    if not iso or len(iso) < 10:
        return None
    y, m, d = iso[:4], iso[5:7], iso[8:10]
    return f"{d}-{m}-{y}"


def load_api_key():
    key = os.environ.get("SETLIST_FM_KEY") or os.environ.get("SETLISTFM_API_KEY")
    if key:
        return key.strip()
    for candidate in (".secrets", ".setlistfm_key", "setlistfm.key"):
        if os.path.exists(candidate):
            with open(candidate) as f:
                line = f.readline().strip()
                if line:
                    return line
    return None


# ============================================================================
# Spreadsheet load (same shape as validate_concert_data.py)
# ============================================================================

def load_concerts(xlsx_path):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Empty spreadsheet")

    header = [str(h).strip() if h else "" for h in rows[0]]

    def col(name_options):
        for i, h in enumerate(header):
            if h.lower() in [n.lower() for n in name_options]:
                return i
        return None

    cd = col(["Date"])
    ca = col(["Headlining Artist", "Artist"])
    cv = col(["Venue"])
    cc = col(["City"])
    cs = col(["State"])
    cacts = col(["Opening Acts"])

    if cd is None or ca is None or cc is None:
        raise ValueError(f"Couldn't find required columns. Found: {header}")

    concerts = []
    for ri, r in enumerate(rows[1:], start=2):
        if not any(v is not None for v in r):
            continue
        artist = str(r[ca]).strip() if r[ca] else None
        if not artist:
            continue
        is_festival = "festival" in artist.lower()
        concerts.append({
            "row_num": ri,
            "date": to_date_str(r[cd]),
            "artist": artist,
            "venue": str(r[cv]).strip() if cv is not None and r[cv] else None,
            "city": str(r[cc]).strip() if cc is not None and r[cc] else None,
            "state": str(r[cs]).strip() if cs is not None and r[cs] else None,
            "opening_acts_raw": str(r[cacts]).strip() if cacts is not None and r[cacts] else None,
            "opening_acts": split_acts(r[cacts] if cacts is not None else None),
            "is_festival": is_festival,
        })
    return concerts


# ============================================================================
# API: query all setlists on a given date in a given city
# ============================================================================

def fetch_setlists_by_city_date(city, date_iso, api_key, max_retries=3):
    """
    Returns: list of setlist dicts from /search/setlists, or {"_error": "..."}.
    Each setlist dict has: artist.name, eventDate, venue.name, venue.city.name, etc.
    Walks paginated results (typically 1-2 pages max for a single city+date).
    """
    sfm_date = iso_to_setlistfm_date(date_iso)
    if not sfm_date:
        return {"_error": "bad date"}

    all_setlists = []
    page = 1
    max_pages = 5  # Safety cap — single-city-single-date should never exceed this

    while page <= max_pages:
        params = urllib.parse.urlencode({
            "cityName": city,
            "date": sfm_date,
            "p": str(page),
        })
        url = f"{API_BASE}/search/setlists?{params}"
        req = urllib.request.Request(url, headers={
            "x-api-key": api_key,
            "Accept": "application/json",
            "User-Agent": "Sethlist/1.0 (personal concert archive)",
        })

        last_error = None
        data = None
        for attempt in range(max_retries):
            try:
                with urllib.request.urlopen(req, timeout=30) as r:
                    data = json.loads(r.read())
                    break
            except urllib.error.HTTPError as e:
                if e.code == 404:
                    return {"_error": "no setlists found", "results": []} if page == 1 else \
                           {"results": all_setlists}
                last_error = f"HTTP {e.code}"
                if e.code in (429, 502, 503, 504) and attempt < max_retries - 1:
                    wait = 2 ** (attempt + 1)
                    print(f"    (got HTTP {e.code}, waiting {wait}s before retry…)",
                          flush=True, file=sys.stderr)
                    time.sleep(wait)
                    continue
                return {"_error": last_error}
            except (TimeoutError, OSError) as e:
                last_error = f"{type(e).__name__}: {e}"
                if attempt < max_retries - 1:
                    wait = 2 ** (attempt + 1)
                    print(f"    (timeout/network error, waiting {wait}s before retry…)",
                          flush=True, file=sys.stderr)
                    time.sleep(wait)
                    continue
                return {"_error": last_error}
            except Exception as e:
                return {"_error": f"{type(e).__name__}: {e}"}

        if data is None:
            return {"_error": last_error or "unknown"}

        page_results = data.get("setlist", []) or []
        all_setlists.extend(page_results)

        # Pagination: stop if we got fewer than `itemsPerPage` results
        items_per_page = data.get("itemsPerPage", 20)
        total = data.get("total", 0)
        if len(all_setlists) >= total or len(page_results) < items_per_page:
            break
        page += 1

    return {"results": all_setlists}


# ============================================================================
# Comparison logic
# ============================================================================

def compare(concert, api_setlists):
    """
    Compare a spreadsheet concert against the setlist.fm setlists at the same
    city+date. Returns a structured comparison result.

    Strategy:
      - Filter API results to those at the same venue (when venue names match
        loosely) OR same date+city if no venue match available.
      - The headliner artist should be among the API results — confirms we're
        looking at the right show. If absent, flag as "show not found on
        setlist.fm".
      - Bands in API list (excluding headliner) are the "API openers".
      - Compare to spreadsheet openers using normalized keys.
    """
    target_artist_key = normalize_artist_key(concert["artist"])
    sheet_opener_keys = {normalize_artist_key(o): o for o in concert["opening_acts"]}

    # Filter API setlists to ones that look like the same show
    # (loose venue name match — sub-string in either direction, case-insensitive)
    venue_lower = (concert["venue"] or "").lower()
    relevant = []
    for sl in api_setlists:
        api_venue = ((sl.get("venue") or {}).get("name") or "").lower()
        if not venue_lower or not api_venue:
            relevant.append(sl)
        elif venue_lower in api_venue or api_venue in venue_lower:
            relevant.append(sl)
        # If neither matches, skip — likely a different show in the same city

    if not relevant:
        # Fall back to all results from same city+date if venue filter eliminates everything
        relevant = api_setlists

    # Build a key->spelling map of every band the API saw at this venue/date
    api_bands = {}  # normalized_key -> display name
    headliner_found = False
    for sl in relevant:
        name = (sl.get("artist") or {}).get("name", "")
        if not name:
            continue
        key = normalize_artist_key(name)
        if key == target_artist_key:
            headliner_found = True
        else:
            # API openers are everyone except the spreadsheet's headliner
            api_bands.setdefault(key, name)

    # Compute set differences
    api_only_keys = set(api_bands.keys()) - set(sheet_opener_keys.keys())
    sheet_only_keys = set(sheet_opener_keys.keys()) - set(api_bands.keys())
    both_keys = set(api_bands.keys()) & set(sheet_opener_keys.keys())

    return {
        "headliner_found_on_api": headliner_found,
        "api_band_count": len(api_bands),
        "api_only": sorted([api_bands[k] for k in api_only_keys]),
        "sheet_only": sorted([sheet_opener_keys[k] for k in sheet_only_keys]),
        "matched": sorted([api_bands[k] for k in both_keys]),
        "raw_relevant_count": len(relevant),
    }


# ============================================================================
# Main flow
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description="Cross-check openers against setlist.fm.")
    parser.add_argument("xlsx", help="Path to Concert History.xlsx")
    parser.add_argument("--out", default="opener_validation_report.md",
                        help="Output report path (default: opener_validation_report.md)")
    parser.add_argument("--cache", default="data/opener_validation_cache.json",
                        help="Cache path for API responses")
    parser.add_argument("--refresh", action="store_true",
                        help="Re-query API even if cached")
    parser.add_argument("--limit", type=int, default=None,
                        help="Validate only the first N shows (smoke test)")
    parser.add_argument("--pause", type=float, default=1.0,
                        help="Seconds between API calls (default: 1.0)")
    args = parser.parse_args()

    api_key = load_api_key()
    if not api_key:
        print("Error: no setlist.fm API key found.", file=sys.stderr)
        print("See prefetch_setlists.py for setup.", file=sys.stderr)
        sys.exit(1)

    if not Path(args.xlsx).exists():
        print(f"Error: {args.xlsx} not found", file=sys.stderr)
        sys.exit(1)

    print(f"Reading {args.xlsx}…", file=sys.stderr)
    concerts = load_concerts(args.xlsx)
    # Filter out festivals + future shows + shows missing required fields
    today = datetime.date.today().strftime("%Y-%m-%d")
    work = []
    skipped_festival = 0
    skipped_future = 0
    skipped_incomplete = 0
    for c in concerts:
        if c["is_festival"]:
            skipped_festival += 1
            continue
        if not c["date"] or c["date"] > today:
            skipped_future += 1
            continue
        if not c["city"]:
            skipped_incomplete += 1
            continue
        # Shows with no openers in the spreadsheet still get checked — the API
        # might confirm that or surface bands the user missed.
        work.append(c)

    if args.limit:
        work = work[:args.limit]

    print(f"  {len(concerts)} total concerts", file=sys.stderr)
    print(f"  {skipped_festival} festivals skipped (per scope)", file=sys.stderr)
    print(f"  {skipped_future} future/undated shows skipped", file=sys.stderr)
    print(f"  {skipped_incomplete} shows missing city skipped", file=sys.stderr)
    print(f"  {len(work)} shows to validate", file=sys.stderr)
    print("", file=sys.stderr)

    # Load API cache
    cache = {}
    cache_path = Path(args.cache)
    if cache_path.exists():
        try:
            cache = json.loads(cache_path.read_text())
        except json.JSONDecodeError:
            print(f"Warning: {args.cache} malformed; starting fresh.", file=sys.stderr)
    cache_path.parent.mkdir(parents=True, exist_ok=True)

    # Walk shows, fetching as needed
    fetched = 0
    cache_hits = 0
    for i, c in enumerate(work, 1):
        cache_key = f"{c['date']}|{c['city']}"
        prefix = f"[{i:>3}/{len(work)}] {c['date']} {c['artist'][:30]:<30}"

        if cache_key in cache and not args.refresh:
            cache_hits += 1
            print(f"{prefix} (cached)", file=sys.stderr)
            continue

        print(f"{prefix} fetching…", file=sys.stderr)
        result = fetch_setlists_by_city_date(c["city"], c["date"], api_key)
        cache[cache_key] = result
        fetched += 1

        # Save incrementally every 10 to avoid losing progress
        if i % 10 == 0 or i == len(work):
            cache_path.write_text(json.dumps(cache, indent=2))
        time.sleep(args.pause)

    cache_path.write_text(json.dumps(cache, indent=2))
    print("", file=sys.stderr)
    print(f"  Fetched: {fetched}, cache hits: {cache_hits}", file=sys.stderr)

    # ----- Build the report -----
    print("Comparing…", file=sys.stderr)

    sections = {
        "exact_match": [],       # spreadsheet & API agree exactly
        "sheet_subset": [],      # all sheet openers found in API; API has more
        "discrepancy": [],       # genuine mismatches in either direction
        "no_api_data": [],       # API returned nothing relevant
        "headliner_missing": [], # API has bands but not the headliner — uncertain
    }

    for c in work:
        cache_key = f"{c['date']}|{c['city']}"
        result = cache.get(cache_key, {})
        if result.get("_error"):
            sections["no_api_data"].append({
                "concert": c, "reason": result["_error"],
            })
            continue
        api_setlists = result.get("results", []) or []
        if not api_setlists:
            sections["no_api_data"].append({
                "concert": c, "reason": "no setlists returned",
            })
            continue

        cmp_result = compare(c, api_setlists)
        bucket = {"concert": c, **cmp_result}

        if not cmp_result["headliner_found_on_api"] and cmp_result["api_band_count"] > 0:
            sections["headliner_missing"].append(bucket)
        elif cmp_result["api_band_count"] == 0:
            sections["no_api_data"].append({
                "concert": c, "reason": "no other bands found at this city/date",
            })
        elif not cmp_result["sheet_only"] and not cmp_result["api_only"]:
            sections["exact_match"].append(bucket)
        elif not cmp_result["sheet_only"]:
            sections["sheet_subset"].append(bucket)
        else:
            sections["discrepancy"].append(bucket)

    # ----- Render markdown -----
    out = []
    out.append("# Opener cross-check report\n")
    out.append(f"_Auto-generated · {len(work)} non-festival shows queried · "
               f"setlist.fm API cross-check._\n")
    out.append(
        "\nFor each show, this compares your `Opening Acts` cell against the bands "
        "setlist.fm logged at the same city on the same date. Mismatches are "
        "candidates to investigate, not auto-corrections.\n"
    )

    # Section: discrepancies (the actionable ones)
    out.append(f"\n## ⚠ Discrepancies ({len(sections['discrepancy'])})\n")
    out.append(
        "Shows where your sheet has openers setlist.fm doesn't have, AND/OR "
        "setlist.fm has bands you don't have. Most actionable — likely typos "
        "or wrong-show entries.\n"
    )
    if not sections["discrepancy"]:
        out.append("\n_No discrepancies found._\n")
    else:
        for d in sections["discrepancy"]:
            c = d["concert"]
            out.append(f"\n### row {c['row_num']} · {c['date']} · `{c['artist']}` @ "
                       f"{c['venue']}, {c['city']}\n")
            if d["sheet_only"]:
                out.append("**In your sheet but not on setlist.fm:**")
                for b in d["sheet_only"]:
                    out.append(f"- `{b}`")
            if d["api_only"]:
                out.append("\n**On setlist.fm but not in your sheet:**")
                for b in d["api_only"]:
                    out.append(f"- `{b}`")
            if d["matched"]:
                out.append(f"\n_Matched: {', '.join(d['matched'])}_")
            out.append("")

    # Section: subset (informational)
    out.append(f"\n## ℹ Sheet is a subset of API ({len(sections['sheet_subset'])})\n")
    out.append(
        "Shows where everything you have is on setlist.fm, but setlist.fm has "
        "additional bands you didn't list. Often this means there were other "
        "openers you don't remember/skipped — informational, not necessarily "
        "errors.\n"
    )
    if not sections["sheet_subset"]:
        out.append("\n_None._\n")
    else:
        for d in sections["sheet_subset"]:
            c = d["concert"]
            out.append(f"\n- **row {c['row_num']}** · {c['date']} · `{c['artist']}` "
                       f"@ {c['venue']}, {c['city']}")
            out.append(f"  - Additional on setlist.fm: {', '.join('`' + b + '`' for b in d['api_only'])}")

    # Section: clean matches (good to confirm)
    out.append(f"\n## ✓ Clean matches ({len(sections['exact_match'])})\n")
    if not sections["exact_match"]:
        out.append("\n_None — your sheet didn't perfectly match the API for any show._\n")
    else:
        out.append("\n_These shows have openers that exactly match what setlist.fm has._\n")
        for d in sections["exact_match"]:
            c = d["concert"]
            note = ", ".join(d["matched"]) if d["matched"] else "(no openers — confirmed solo show)"
            out.append(f"\n- **row {c['row_num']}** · {c['date']} · `{c['artist']}` "
                       f"@ {c['venue']}, {c['city']} — {note}")

    # Section: headliner missing (uncertain)
    out.append(f"\n## ❓ Headliner not on setlist.fm at this venue/date "
               f"({len(sections['headliner_missing'])})\n")
    out.append(
        "Setlist.fm has bands logged for this city/date, but the headliner you "
        "listed isn't among them. Could mean: wrong city/date in the sheet, "
        "headliner spelled differently in setlist.fm, or two unrelated shows in "
        "the same city that night. These reports are uncertain — review by hand.\n"
    )
    if not sections["headliner_missing"]:
        out.append("\n_None._\n")
    else:
        for d in sections["headliner_missing"]:
            c = d["concert"]
            out.append(f"\n- **row {c['row_num']}** · {c['date']} · `{c['artist']}` "
                       f"@ {c['venue']}, {c['city']}")
            other_bands = sorted(set(d["api_only"] + d["matched"]))
            out.append(f"  - Bands setlist.fm has at this city/date: {', '.join('`'+b+'`' for b in other_bands)}")

    # Section: no API data (informational)
    out.append(f"\n## — No setlist.fm data ({len(sections['no_api_data'])})\n")
    out.append(
        "Setlist.fm doesn't have any setlists for these shows. Common for older "
        "small-club shows or smaller bands. Not actionable.\n"
    )
    if sections["no_api_data"]:
        for d in sections["no_api_data"]:
            c = d["concert"]
            out.append(f"- row {c['row_num']} · {c['date']} · `{c['artist']}` "
                       f"@ {c['city']} — _{d['reason']}_")

    out.append("\n---\n")
    out.append(
        f"_Cache: `{args.cache}`. Re-run with `--refresh` to re-query the API. "
        f"Edit the spreadsheet, re-run `build_data.py`, then re-run this script "
        f"(uses cache by default; will be fast)._\n"
    )

    Path(args.out).write_text("\n".join(out))
    print(f"Report written to {args.out}", file=sys.stderr)
    print("", file=sys.stderr)
    print(f"Summary:", file=sys.stderr)
    print(f"  ⚠ Discrepancies:           {len(sections['discrepancy'])}", file=sys.stderr)
    print(f"  ℹ Sheet is a subset:       {len(sections['sheet_subset'])}", file=sys.stderr)
    print(f"  ✓ Clean matches:           {len(sections['exact_match'])}", file=sys.stderr)
    print(f"  ❓ Headliner missing:       {len(sections['headliner_missing'])}", file=sys.stderr)
    print(f"  — No API data:             {len(sections['no_api_data'])}", file=sys.stderr)


if __name__ == "__main__":
    main()
