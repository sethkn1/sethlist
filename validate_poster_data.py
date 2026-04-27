"""
validate_poster_data.py — Validation passes for the Poster Collection xlsx.

Five checks, all internal-consistency (no API calls):

1. SPELLING — Group every Artist (band) name by normalized key. Within each
   cluster, multiple distinct spellings indicate a typo or inconsistency.
   Same logic as validate_concert_data.py.

2. CROSS-FILE BAND NAMES — Compare the poster Artist column against the
   Concert History headliner column. If the same band's posters use a
   different spelling than the concerts (e.g., "NIN" in posters but
   "Nine Inch Nails" in concerts), flag for normalization.

3. ORPHAN POSTERS — Posters for shows that don't appear in the concerts
   spreadsheet. Could be: (a) a show you have a poster for but didn't
   attend, (b) a typo in date/artist/location, (c) a missing concert row.

4. ATTENDED FLAG VS CONCERTS — A poster marked "Attended: Yes" should
   correspond to a concert row. Mismatches (poster says attended but no
   matching concert; or marked No but a concert exists) are worth review.

5. STRUCTURAL — Rows with formatting / data-entry issues:
   - Missing required fields (Date, Artist, Location)
   - Number field formatted oddly (e.g., "47/" or "/50" or non-fraction
     where one is expected)
   - Notes field that mentions a different venue/date than the row's own
   - Duplicate poster rows (same date+artist+illustrator+number)

Usage:
    python3 validate_poster_data.py "Poster Collection.xlsx" "Concert History.xlsx"
    python3 validate_poster_data.py "Poster Collection.xlsx" "Concert History.xlsx" --out report.md

If you skip the concerts xlsx path, checks 2/3/4 are skipped and only the
internal-only checks (1 and 5) run.
"""

import argparse
import datetime
import re
import sys
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ============================================================================
# Shared helpers (mirroring app.js + other validators)
# ============================================================================

def normalize_artist_key(s):
    if not s:
        return ""
    s = str(s).lower()
    s = re.sub(r"^the\s+", "", s)
    s = re.sub(r"[^a-z0-9]", "", s)
    return s


def split_acts(s):
    if not s:
        return []
    parts = re.split(r"[,;]", str(s))
    out = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        p = re.sub(r"\s*\([^)]*\)\s*", "", p).strip()
        if p:
            out.append(p)
    return out


def to_date_str(v):
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, datetime.date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, (int, float)):
        return (datetime.date(1899, 12, 30) + datetime.timedelta(days=int(v))).strftime("%Y-%m-%d")
    s = str(v).strip()
    return s if s else None


def col_index(header, name_options):
    """Find a column index by trying multiple header names (case-insensitive)."""
    options_lower = [n.lower() for n in name_options]
    for i, h in enumerate(header):
        if h and h.lower().strip() in options_lower:
            return i
    return None


# ============================================================================
# Loaders
# ============================================================================

def load_posters(xlsx_path):
    """Read the 'Poster List' sheet from the poster collection xlsx."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Poster List" not in wb.sheetnames:
        raise ValueError(f"Sheet 'Poster List' not found. Sheets: {wb.sheetnames}")
    ws = wb["Poster List"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Poster List sheet is empty")

    header = [str(h).strip() if h else "" for h in rows[0]]

    cd = col_index(header, ["Date"])
    ca = col_index(header, ["Artist"])
    cill = col_index(header, ["Artist/Illustrator", "Illustrator"])
    cl = col_index(header, ["Location", "City"])
    ct = col_index(header, ["Type"])
    cv = col_index(header, ["Variant"])
    cnum = col_index(header, ["Number"])
    cscope = col_index(header, ["Tour/Show Specific", "Scope"])
    cauto = col_index(header, ["Autographed"])
    cframe = col_index(header, ["Framed"])
    catt = col_index(header, ["Attended"])
    cnotes = col_index(header, ["Poster Notes:", "Poster Notes", "Notes"])

    if cd is None or ca is None or cl is None:
        raise ValueError(f"Couldn't find required columns. Found: {header}")

    posters = []
    for ri, r in enumerate(rows[1:], start=2):
        if not any(v is not None for v in r):
            continue
        artist = str(r[ca]).strip() if r[ca] else None
        if not artist:
            continue
        posters.append({
            "row_num": ri,
            "date": to_date_str(r[cd]),
            "artist": artist,
            "illustrator": str(r[cill]).strip() if cill is not None and r[cill] else None,
            "location": str(r[cl]).strip() if cl is not None and r[cl] else None,
            "type": str(r[ct]).strip() if ct is not None and r[ct] else None,
            "variant": str(r[cv]).strip() if cv is not None and r[cv] else None,
            "number": str(r[cnum]).strip() if cnum is not None and r[cnum] else None,
            "scope": str(r[cscope]).strip() if cscope is not None and r[cscope] else None,
            "autographed": str(r[cauto]).strip() if cauto is not None and r[cauto] else None,
            "framed": str(r[cframe]).strip() if cframe is not None and r[cframe] else None,
            "attended": str(r[catt]).strip() if catt is not None and r[catt] else None,
            "notes": str(r[cnotes]).strip() if cnotes is not None and r[cnotes] else None,
            "is_festival": "festival" in artist.lower(),
        })
    return posters


def load_concerts(xlsx_path):
    """Read the concert history xlsx — same logic as validate_concert_data.py."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Empty concerts sheet")

    header = [str(h).strip() if h else "" for h in rows[0]]
    cd = col_index(header, ["Date"])
    ca = col_index(header, ["Headlining Artist", "Artist"])
    cc = col_index(header, ["City"])
    cv = col_index(header, ["Venue"])
    cacts = col_index(header, ["Opening Acts"])

    if cd is None or ca is None:
        raise ValueError(f"Couldn't find required columns. Found: {header}")

    concerts = []
    for ri, r in enumerate(rows[1:], start=2):
        if not any(v is not None for v in r):
            continue
        artist = str(r[ca]).strip() if r[ca] else None
        if not artist:
            continue
        concerts.append({
            "row_num": ri,
            "date": to_date_str(r[cd]),
            "artist": artist,
            "city": str(r[cc]).strip() if cc is not None and r[cc] else None,
            "venue": str(r[cv]).strip() if cv is not None and r[cv] else None,
            "opening_acts": split_acts(r[cacts] if cacts is not None else None),
            "is_festival": "festival" in artist.lower(),
        })
    return concerts


# ============================================================================
# Pass 1: Spelling clusters within posters
# ============================================================================

def find_spelling_clusters(posters):
    """Same idea as validate_concert_data.py — group artists by normalized key."""
    spelling_appearances = defaultdict(list)
    spelling_key = {}
    for p in posters:
        sp = p["artist"]
        spelling_appearances[sp].append(p)
        spelling_key[sp] = normalize_artist_key(sp)

    by_key = defaultdict(list)
    for sp, key in spelling_key.items():
        if key:
            by_key[key].append(sp)

    clusters = []
    for key, spellings in by_key.items():
        if len(spellings) <= 1:
            continue
        counts = sorted(
            [(sp, len(spelling_appearances[sp]), spelling_appearances[sp])
             for sp in spellings],
            key=lambda x: (-x[1], x[0])
        )
        clusters.append({
            "key": key,
            "spellings": counts,
            "canonical_guess": counts[0][0],
        })
    clusters.sort(key=lambda c: -sum(n for _, n, _ in c["spellings"]))
    return clusters


# ============================================================================
# Pass 2: Cross-file band-name inconsistencies
# ============================================================================

def find_cross_file_band_mismatches(posters, concerts):
    """
    For each unique band that appears as a poster Artist, check whether the
    same band appears in the concert headliners (or opening acts) under a
    different exact spelling. We use normalized keys to match.

    Reports cases where the spelling differs, suggesting standardization.
    """
    # Build poster-side spelling map
    poster_spellings = defaultdict(set)  # normalized key -> set of poster spellings
    for p in posters:
        if p["is_festival"]:
            continue  # festivals are handled separately
        k = normalize_artist_key(p["artist"])
        if k:
            poster_spellings[k].add(p["artist"])

    # Build concert-side spelling map (headliners + openers)
    concert_spellings = defaultdict(set)
    for c in concerts:
        if c["artist"] and not c["is_festival"]:
            k = normalize_artist_key(c["artist"])
            if k:
                concert_spellings[k].add(c["artist"])
        for opener in c["opening_acts"]:
            k = normalize_artist_key(opener)
            if k:
                concert_spellings[k].add(opener)

    mismatches = []
    for k, p_spellings in poster_spellings.items():
        c_spellings = concert_spellings.get(k, set())
        if not c_spellings:
            continue  # poster band not in any concert — handled by orphan check
        # Are the spelling sets identical?
        if p_spellings != c_spellings:
            extra_in_posters = p_spellings - c_spellings
            extra_in_concerts = c_spellings - p_spellings
            if extra_in_posters or extra_in_concerts:
                mismatches.append({
                    "key": k,
                    "in_posters": sorted(p_spellings),
                    "in_concerts": sorted(c_spellings),
                    "extra_in_posters": sorted(extra_in_posters),
                    "extra_in_concerts": sorted(extra_in_concerts),
                })
    mismatches.sort(key=lambda m: m["key"])
    return mismatches


# ============================================================================
# Pass 3: Orphan posters (no matching concert)
# ============================================================================

def find_orphan_posters(posters, concerts):
    """
    A poster is "orphan" if there's no concert with the same date AND a
    matching artist (by normalized key, in either headliner OR opener slot).

    Returns posters that don't have a corresponding concert.
    """
    # Build (date, normalized-band-key) → list of concerts
    concert_index = defaultdict(list)
    for c in concerts:
        if not c["date"]:
            continue
        if c["artist"]:
            concert_index[(c["date"], normalize_artist_key(c["artist"]))].append(c)
        # Festival rows: also index by their festival "name" (which is in c.artist)
        # so a festival poster can match the festival concert row.
        for opener in c["opening_acts"]:
            concert_index[(c["date"], normalize_artist_key(opener))].append(c)

    orphans = []
    for p in posters:
        if not p["date"] or not p["artist"]:
            continue
        key = (p["date"], normalize_artist_key(p["artist"]))
        if key not in concert_index:
            orphans.append(p)
    return orphans


# ============================================================================
# Pass 4: Attended flag vs concerts
# ============================================================================

def find_attended_mismatches(posters, concerts):
    """
    A poster marked Attended=Yes should match a concert; marked No means you
    don't have an attended-show record. Cross-check the flag against actual
    concert presence.

    Returns:
      attended_yes_no_concert: poster says yes but no concert row exists
      attended_no_with_concert: poster says no but a concert row matches
    """
    concert_dates_by_band = defaultdict(set)  # band-key -> set of concert dates
    for c in concerts:
        if not c["date"]:
            continue
        if c["artist"]:
            concert_dates_by_band[normalize_artist_key(c["artist"])].add(c["date"])
        for opener in c["opening_acts"]:
            concert_dates_by_band[normalize_artist_key(opener)].add(c["date"])

    yes_no_concert = []
    no_with_concert = []
    for p in posters:
        if not p["date"] or not p["artist"]:
            continue
        attended = (p["attended"] or "").lower()
        bk = normalize_artist_key(p["artist"])
        has_concert = p["date"] in concert_dates_by_band.get(bk, set())

        if attended == "yes" and not has_concert:
            yes_no_concert.append(p)
        elif attended == "no" and has_concert:
            no_with_concert.append(p)
    return yes_no_concert, no_with_concert


# ============================================================================
# Pass 5: Structural / data-entry concerns
# ============================================================================

def find_structural_issues(posters):
    """Per-row formatting checks."""
    issues = []
    by_signature = defaultdict(list)  # (date, artist, illustrator, number) → rows

    for p in posters:
        row_issues = []

        # Required fields
        if not p["date"]:
            row_issues.append("missing date")
        if not p["artist"]:
            row_issues.append("missing artist")
        if not p["location"]:
            row_issues.append("missing location")

        # Number formatting
        num = p["number"]
        if num and num.lower() not in ("unnumbered", "ap", "n/a", "none"):
            # Expected: "47/100" or "47/100 + AP" or just a fraction
            # Flag if it doesn't have a slash, or has a trailing/leading slash
            if "/" not in num:
                row_issues.append(f"`Number` = {num!r} but doesn't look like a fraction")
            elif num.endswith("/") or num.startswith("/"):
                row_issues.append(f"`Number` = {num!r} has a stray slash")

        # Boolean-ish fields
        for field in ("autographed", "framed", "attended"):
            v = p[field]
            if v and v.lower() not in ("yes", "no", "y", "n", "true", "false"):
                row_issues.append(f"`{field.capitalize()}` = {v!r} (expected Yes/No)")

        # Notes mentioning a different venue or date than the row's own
        # (best-effort: simple substring checks against common venue keywords)
        notes_lower = (p["notes"] or "").lower()
        location_lower = (p["location"] or "").lower()
        # Look for venue mentions in notes that don't match the location
        # We only flag obvious cases: city names of major venues that differ
        for venue_hint in ["red rocks", "fenway", "barclays", "msg",
                           "madison square", "dodger", "wrigley"]:
            if venue_hint in notes_lower and venue_hint not in location_lower:
                # Only flag if the location field clearly disagrees
                row_issues.append(
                    f"notes mention `{venue_hint}` but Location is `{p['location']}`"
                )
                break

        # Build dedup signature
        sig = (
            p["date"] or "",
            normalize_artist_key(p["artist"] or ""),
            (p["illustrator"] or "").strip().lower(),
            (p["number"] or "").strip().lower(),
        )
        # Only consider it a real dup if all four fields are present and informative
        if all(s for s in sig) and sig[3] not in ("unnumbered", "n/a"):
            by_signature[sig].append(p)

        if row_issues:
            issues.append({"poster": p, "issues": row_issues})

    # Dedup signatures with >1 entry
    duplicates = []
    for sig, group in by_signature.items():
        if len(group) > 1:
            duplicates.append({"signature": sig, "rows": group})

    return issues, duplicates


# ============================================================================
# Report rendering
# ============================================================================

def render_report(posters, concerts, results):
    out = []
    out.append("# Poster collection validation report\n")
    cs_part = (f", cross-checked against {len(concerts)} concerts"
               if concerts is not None else "")
    out.append(f"_Auto-generated · {len(posters)} posters scanned{cs_part}._\n")

    # ---- 1. Spelling clusters ----
    out.append("\n## 1. Spelling clusters within posters\n")
    out.append("Bands whose name appears with multiple spellings in the poster sheet.\n")
    clusters = results["spelling_clusters"]
    if not clusters:
        out.append("\n_No clusters — every band's name is consistent within posters._\n")
    else:
        out.append(f"\n**{len(clusters)} cluster(s):**\n")
        for c in clusters:
            out.append(f"\n### `{c['canonical_guess']}` cluster\n")
            for sp, n, apps in c["spellings"]:
                marker = "  ✓ canonical" if sp == c["canonical_guess"] else "  ⚠ suspect"
                out.append(f"- **`{sp}`** — {n} poster{'s' if n != 1 else ''}{marker}")
                for app in apps[:3]:
                    out.append(f"    - row {app['row_num']} · {app['date']} · "
                               f"`{app['location']}` · type=`{app['type']}`")
                if len(apps) > 3:
                    out.append(f"    - …and {len(apps)-3} more")

    # ---- 2. Cross-file band mismatches ----
    if concerts is None:
        out.append("\n## 2. Cross-file band-name comparison\n")
        out.append("\n_Skipped — no concerts file provided._\n")
    else:
        out.append("\n## 2. Band-name mismatches vs Concert History\n")
        out.append(
            "Same band appears in both files but with different spellings. "
            "Recommended: standardize on one spelling (typically the one "
            "concerts uses, since that's the canonical headliner display).\n"
        )
        mm = results["cross_file_mismatches"]
        if not mm:
            out.append("\n_No cross-file mismatches — all band names align._\n")
        else:
            out.append(f"\n**{len(mm)} band(s) with cross-file inconsistencies:**\n")
            for m in mm:
                out.append(f"\n- normalized key `{m['key']}`")
                out.append(f"  - in posters: {', '.join('`'+s+'`' for s in m['in_posters'])}")
                out.append(f"  - in concerts: {', '.join('`'+s+'`' for s in m['in_concerts'])}")
                if m["extra_in_posters"]:
                    out.append(f"  - ⚠ posters use spelling(s) NOT in concerts: "
                               f"{', '.join('`'+s+'`' for s in m['extra_in_posters'])}")

    # ---- 3. Orphan posters ----
    if concerts is None:
        out.append("\n## 3. Orphan posters\n")
        out.append("\n_Skipped — no concerts file provided._\n")
    else:
        out.append("\n## 3. Orphan posters (no matching concert row)\n")
        out.append(
            "Posters whose `Date + Artist` doesn't match any concert in the "
            "Concert History sheet. These could be:\n"
            "- Posters for shows you didn't attend (if `Attended = No`, that's expected)\n"
            "- Typos in date or artist that broke the link\n"
            "- Missing concert rows\n"
        )
        orphans = results["orphans"]
        if not orphans:
            out.append("\n_No orphans — every poster matches a concert row._\n")
        else:
            out.append(f"\n**{len(orphans)} poster(s) with no matching concert:**\n")
            for p in orphans:
                attended_note = f" · attended=`{p['attended']}`" if p["attended"] else ""
                out.append(f"\n- **row {p['row_num']}** · {p['date']} · `{p['artist']}` "
                           f"@ {p['location']}{attended_note}")
                if p["notes"]:
                    out.append(f"  - notes: _{p['notes']}_")

    # ---- 4. Attended mismatches ----
    if concerts is None:
        out.append("\n## 4. Attended-flag cross-check\n")
        out.append("\n_Skipped — no concerts file provided._\n")
    else:
        out.append("\n## 4. Attended-flag mismatches\n")
        out.append(
            "Where the poster's `Attended` flag disagrees with whether you "
            "have a matching concert record. Most actionable: posters marked "
            "`Yes` for which there's no concert (likely a date/artist typo).\n"
        )
        yes_no_concert = results["attended_yes_no_concert"]
        no_with_concert = results["attended_no_with_concert"]
        if not yes_no_concert and not no_with_concert:
            out.append("\n_No mismatches._\n")
        else:
            if yes_no_concert:
                out.append(f"\n**Marked `Attended=Yes` but no concert found ({len(yes_no_concert)}):**\n")
                for p in yes_no_concert:
                    out.append(f"\n- **row {p['row_num']}** · {p['date']} · `{p['artist']}` "
                               f"@ {p['location']}")
                    if p["notes"]:
                        out.append(f"  - notes: _{p['notes']}_")
            if no_with_concert:
                out.append(f"\n**Marked `Attended=No` but a concert exists ({len(no_with_concert)}):**\n")
                for p in no_with_concert:
                    out.append(f"\n- **row {p['row_num']}** · {p['date']} · `{p['artist']}` "
                               f"@ {p['location']}")
                    if p["notes"]:
                        out.append(f"  - notes: _{p['notes']}_")

    # ---- 5. Structural ----
    out.append("\n## 5. Structural / data-entry concerns\n")
    issues = results["structural_issues"]
    duplicates = results["duplicates"]

    if not issues and not duplicates:
        out.append("\n_No structural issues._\n")
    else:
        if issues:
            out.append(f"\n**{len(issues)} row(s) with formatting issues:**\n")
            for i in issues:
                p = i["poster"]
                out.append(f"\n- **row {p['row_num']}** · {p['date']} · "
                           f"`{p['artist']}` @ {p['location']}")
                for issue in i["issues"]:
                    out.append(f"  - ⚠ {issue}")
        if duplicates:
            out.append(f"\n**{len(duplicates)} duplicate group(s) "
                       f"(same date+artist+illustrator+number):**\n")
            for d in duplicates:
                sig = d["signature"]
                out.append(f"\n- key: date=`{sig[0]}`, illustrator=`{sig[2]}`, "
                           f"number=`{sig[3]}`")
                for p in d["rows"]:
                    out.append(f"  - row {p['row_num']} · type=`{p['type']}` · "
                               f"variant=`{p['variant']}`")

    out.append("\n---\n")
    out.append("_End of report._\n")
    return "\n".join(out)


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description="Validate poster collection data.")
    parser.add_argument("posters_xlsx", help="Path to Poster Collection.xlsx")
    parser.add_argument("concerts_xlsx", nargs="?", default=None,
                        help="Optional path to Concert History.xlsx for cross-check")
    parser.add_argument("--out", default=None, help="Output report path (default: stdout)")
    args = parser.parse_args()

    if not Path(args.posters_xlsx).exists():
        print(f"Error: {args.posters_xlsx} not found", file=sys.stderr)
        sys.exit(1)

    print(f"Reading {args.posters_xlsx}…", file=sys.stderr)
    posters = load_posters(args.posters_xlsx)
    print(f"  {len(posters)} posters loaded", file=sys.stderr)

    concerts = None
    if args.concerts_xlsx:
        if not Path(args.concerts_xlsx).exists():
            print(f"Error: {args.concerts_xlsx} not found", file=sys.stderr)
            sys.exit(1)
        print(f"Reading {args.concerts_xlsx}…", file=sys.stderr)
        concerts = load_concerts(args.concerts_xlsx)
        print(f"  {len(concerts)} concerts loaded", file=sys.stderr)

    results = {}

    print("Pass 1: spelling clusters…", file=sys.stderr)
    results["spelling_clusters"] = find_spelling_clusters(posters)
    print(f"  {len(results['spelling_clusters'])} cluster(s)", file=sys.stderr)

    if concerts is not None:
        print("Pass 2: cross-file band mismatches…", file=sys.stderr)
        results["cross_file_mismatches"] = find_cross_file_band_mismatches(posters, concerts)
        print(f"  {len(results['cross_file_mismatches'])} mismatch(es)", file=sys.stderr)

        print("Pass 3: orphan posters…", file=sys.stderr)
        results["orphans"] = find_orphan_posters(posters, concerts)
        print(f"  {len(results['orphans'])} orphan(s)", file=sys.stderr)

        print("Pass 4: attended-flag cross-check…", file=sys.stderr)
        yes_nc, no_wc = find_attended_mismatches(posters, concerts)
        results["attended_yes_no_concert"] = yes_nc
        results["attended_no_with_concert"] = no_wc
        print(f"  Yes-but-no-concert: {len(yes_nc)}, No-but-concert-exists: {len(no_wc)}",
              file=sys.stderr)

    print("Pass 5: structural issues…", file=sys.stderr)
    issues, dups = find_structural_issues(posters)
    results["structural_issues"] = issues
    results["duplicates"] = dups
    print(f"  {len(issues)} row(s) with issues, {len(dups)} duplicate group(s)",
          file=sys.stderr)

    report = render_report(posters, concerts, results)
    if args.out:
        Path(args.out).write_text(report)
        print(f"\nReport written to {args.out}", file=sys.stderr)
    else:
        print(report)


if __name__ == "__main__":
    main()
